"""
主要说明, 该脚本和excel_to_json是有关系的

excel_to_json是根据 一个excel导出一个json结构, 保证excle的顺序
A B C
A D E
F G H
导出的结构式
[
    {
        "name": "A",
        "children": [
            {
                "name": "B",
                "children": [
                    {"name": "C"}
                ]
            },
            {
                "name": "D",
                "children":[
                    {"name": "E"}
                ]
            }
        ]
    }

本脚本主要针对txt文件
1-1 1 2 3 4 5 6 7 8
1-2 3 4 5 6 7 8 9 0
1-3 2 3 4 5 6 7 8 9
2-1 2....
2-2 3....
2-3 ....
1-1指知识树的末节点的第几个知识点的难易程度, 后面的数字是推荐的题目
这样的格式进入到excel里
数的认识|易  数的认识|中 数的认识|难
1           4         7
2           5         8
3           6         9

脚本的执行格式
python write_excel.py xxx.txt source.xlsx aim.xlsx
"""

import sys


with open(sys.argv[1], 'r', encoding='utf-8') as fp:
    import xlrd
    from openpyxl import Workbook
    level_map = {
        '1': '|易',
        '2': '|中',
        '3': '|难'
    }
    xls = xlrd.open_workbook(sys.argv[2])
    sheet_name = xls.sheet_names()[0]
    sheet = xls.sheet_by_name(sheet_name)
    cols_val = sheet.col_values(sheet.ncols-1)

    wb = Workbook()
    sheet = wb.get_active_sheet()
    sheet.title="Sheet1"
    for index, line in enumerate(fp.readlines()):
        datas = line.split(' ')
        knowledge_index, level = datas[0].split('-')
        name = cols_val[int(knowledge_index)-1]+level_map[level]
        print(name)
        sheet.cell(row=1, column=index+1).value=name
        for _index, question_id in enumerate(datas[1:]):
            sheet.cell(row=_index+2, column=index+1).value=question_id.strip()
    wb.save(sys.argv[3])